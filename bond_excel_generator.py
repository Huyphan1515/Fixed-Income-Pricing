import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import quote_sheetname
from openpyxl.styles import Font
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from typing import List, Tuple, Dict, Any

def get_coupon_schedule(
    bond_type: str,
    issue_date: datetime,
    maturity_date: datetime,
    face_value: float,
    coupon_rate: float,
    frequency: int,
    coupon_rates: List[float],
    num_periods: int
) -> Tuple[List[datetime], List[List[Any]]]:
    coupon_dates = []
    cashflows = []
    temp_date = issue_date

    if bond_type == "zero" or num_periods == 0 or not coupon_rates:
        # Zero coupon: just one payment at maturity
        coupon_dates = [maturity_date]
        cashflows = [[maturity_date.date(), '', '', face_value]]
        return coupon_dates, cashflows

    months_per = (12 // frequency) if frequency else 12
    for i in range(num_periods):
        next_date = temp_date + relativedelta(months=months_per)
        if next_date > maturity_date or i == num_periods - 1:
            next_date = maturity_date
        coupon_dates.append(next_date)
        ex_cd = next_date - timedelta(days=10)
        prev_cd = temp_date
        year_frac = (next_date - prev_cd).days / 365
        rate = coupon_rates[i] if i < len(coupon_rates) else coupon_rate
        cf = face_value * (rate / 100) * year_frac
        if i == num_periods - 1:
            cf += face_value
        cashflows.append([next_date.date(), ex_cd.date(), f"{rate}%", round(cf, 4)])
        temp_date = next_date
    return coupon_dates, cashflows

def generate_excel(
    bond_type: str,
    issue_date: str,
    maturity_date: str,
    face_value: float,
    bought_date: str,
    sold_date: str,
    quantity: int,
    client_type: str,
    product_type: str,
    trading_fee: float,
    apply_trading_fee: bool,
    num_periods: int,
    coupon_rates: List[float],
    discount_rate: float,
    coupon_rate: float,
    frequency: int,
    filepath: str
) -> Tuple[float, float, Dict[str, Any]]:
    issue_date_dt = datetime.strptime(issue_date, "%Y-%m-%d")
    maturity_date_dt = datetime.strptime(maturity_date, "%Y-%m-%d")
    bought_dt = datetime.strptime(bought_date, "%Y-%m-%d")
    sold_dt = datetime.strptime(sold_date, "%Y-%m-%d")

    coupon_dates, cashflows = get_coupon_schedule(
        bond_type, issue_date_dt, maturity_date_dt, face_value, coupon_rate, frequency, coupon_rates, num_periods
    )

    cf_df = pd.DataFrame(cashflows, columns=["Coupon Date", "Ex-Coupon Date", "Coupon Rate", "Cashflow"])

    wb = Workbook()
    ws_cf = wb.active
    ws_cf.title = "Cash Flow Table"
    for r in dataframe_to_rows(cf_df, index=False, header=True):
        ws_cf.append(r)
    for cell in ws_cf[1]:
        cell.font = Font(bold=True)

    ws_input = wb.create_sheet("User Input")
    ws_input.append(["Bought Date", "Sold Date", "Quantities", "Client Type", "Trading Fee"])
    ws_input.append([bought_date, sold_date, quantity, client_type, trading_fee / 100])

    ws_tax = wb.create_sheet("Tax Table")
    ws_tax.append(["Client Type", "Coupon Tax", "Transaction Tax"])
    ws_tax.append(["Individual", 0.05, 0.001])
    ws_tax.append(["Corporation", 0.0, 0.0])

    ws_pv = wb.create_sheet("PV Table")
    ws_pv.append([
        "Coupon Date", "Ex-Coupon Date", "Coupon Rate", "Cashflow", "Year Frac", "Discount Factor", "Present Value"
    ])
    for cell in ws_pv[1]:
        cell.font = Font(bold=True)

    cf_sheet = quote_sheetname("Cash Flow Table")
    input_sheet = quote_sheetname("User Input")
    for i in range(len(coupon_dates)):
        row = i + 2
        ws_pv[f"A{row}"].value = f"={cf_sheet}!A{row}"
        ws_pv[f"B{row}"].value = f"={cf_sheet}!B{row}"
        ws_pv[f"C{row}"].value = f"={cf_sheet}!C{row}"
        ws_pv[f"D{row}"].value = f"={cf_sheet}!D{row}"
        ws_pv[f"E{row}"].value = ""  # Can be filled in backend if desired
        ws_pv[f"F{row}"].value = ""
        ws_pv[f"G{row}"].value = ""

    # Tax logic
    txn_tax = 0.001 if client_type == "Individual" else 0.0
    coupon_tax = 0.05 if client_type == "Individual" else 0.0
    trading_fee_decimal = trading_fee / 100
    received_cashflows = [cf for cf, cd in zip(cashflows, coupon_dates) if bought_dt <= cd <= sold_dt]
    total_coupon_received = sum(cf[3] * (1 - coupon_tax) for cf in received_cashflows)

    # Price logic
    if bond_type == "zero":
        buy_price = face_value / ((1 + discount_rate / 100) ** ((maturity_date_dt - bought_dt).days / 365))
        sell_price = face_value / ((1 + discount_rate / 100) ** ((maturity_date_dt - sold_dt).days / 365))
    else:
        buy_price = sum([
            cf[3] / ((1 + discount_rate / 100) ** ((cd - bought_dt).days / 365))
            for cf, cd in zip(cashflows, coupon_dates)
        ])
        sell_price = sum([
            cf[3] / ((1 + discount_rate / 100) ** ((cd - sold_dt).days / 365)) * (1 - txn_tax - trading_fee_decimal)
            for cf, cd in zip(cashflows, coupon_dates) if cd > sold_dt
        ])

    ws_summary = wb.create_sheet("Investment Summary")
    ws_summary.append(["Buy Price", round(buy_price * (1 + trading_fee_decimal), 4)])
    ws_summary.append(["Transaction Tax Rate", txn_tax])
    ws_summary.append(["Trading Fee", trading_fee_decimal])
    ws_summary.append(["Total Coupon Received", round(total_coupon_received * (1 - coupon_tax), 4)])
    ws_summary.append(["Sell Price", round(sell_price * (1 - txn_tax), 4)])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)

    ws_table = wb.create_sheet("Investment Table")
    ws_table.append(["Date", "Event", "Net Amount Per Bond"])
    ws_table.append([bought_date, "Buy Bond", round(buy_price * (1 + trading_fee_decimal), 4)])
    for cf, cd in zip(cashflows, coupon_dates):
        if bought_dt <= cd <= sold_dt:
            ws_table.append([cd, "Coupon Received", round(cf[3] * (1 - coupon_tax), 4)])
    ws_table.append([sold_date, "Sell Bond", round(sell_price * (1 - txn_tax), 4)])
    for cell in ws_table[1]:
        cell.font = Font(bold=True)

    wb.save(filepath)

    summary = {
        "buy_price": round(buy_price * quantity, 4),
        "sell_price": round(sell_price * quantity * (1 - txn_tax), 4),
        "coupon_received": round(total_coupon_received * quantity, 4),
        "txn_tax": txn_tax,
        "trading_fee": trading_fee_decimal,
        "investment_table": [
            {"date": bought_date, "event": "Buy Bond", "amount": round(buy_price * (1 + trading_fee_decimal), 4)},
            *[
                {"date": str(cf[0]), "event": "Coupon Received", "amount": round(cf[3] * (1 - coupon_tax), 4)}
                for cf, cd in zip(cashflows, coupon_dates) if bought_dt <= cd <= sold_dt
            ],
            {"date": sold_date, "event": "Sell Bond", "amount": round(sell_price * (1 - txn_tax), 4)}
        ]
    }
    return buy_price * quantity, sell_price * quantity, summary
